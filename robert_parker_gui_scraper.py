import asyncio
import json
import os
import threading
import time
from datetime import datetime
from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font
import tkinter as tk
from tkinter import ttk, scrolledtext, messagebox, filedialog
from asyncio import Semaphore
import re

class RobertParkerScraper:
    def __init__(self, email, password, max_concurrent=5, requests_per_minute=30):
        self.email = email
        self.password = password
        self.cookies_file = "robert_parker_cookies.json"
        self.max_concurrent = max_concurrent
        self.requests_per_minute = requests_per_minute
        self.semaphore = Semaphore(max_concurrent)
        self.rate_limit_delay = 60.0 / requests_per_minute
        self.last_request_time = 0
        
    async def rate_limit(self):
        """Ensure we don't exceed the rate limit"""
        current_time = time.time()
        time_since_last = current_time - self.last_request_time
        if time_since_last < self.rate_limit_delay:
            await asyncio.sleep(self.rate_limit_delay - time_since_last)
        self.last_request_time = time.time()

    async def setup_browser(self):
        """Set up Playwright browser with persistent context and improved configuration"""
        self.playwright = await async_playwright().start()
        
        # Use persistent context for better performance and cookie management
        user_data_dir = os.path.join(os.getcwd(), "browser_data")
        self.browser = await self.playwright.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            headless=False,  # Set to True for production
            args=[
                '--no-sandbox',
                '--disable-dev-shm-usage',
                '--disable-gpu',
                '--disable-web-security',
                '--disable-features=VizDisplayCompositor',
                '--disable-background-timer-throttling',
                '--disable-backgrounding-occluded-windows',
                '--disable-renderer-backgrounding',
                '--disable-ipc-flooding-protection',
                '--disable-default-apps',
                '--disable-extensions',
                '--disable-plugins',
                '--disable-sync',
                '--disable-translate',
                '--no-first-run',
                '--no-default-browser-check',
                '--disable-background-networking',
                '--disable-component-extensions-with-background-pages',
                '--disable-client-side-phishing-detection',
                '--disable-hang-monitor',
                '--disable-prompt-on-repost',
                '--disable-domain-reliability',
                '--disable-features=TranslateUI',
                '--disable-print-preview',
                '--disable-save-password-bubble',
                '--disable-single-click-autofill',
                '--disable-spellcheck-api',
                '--disable-threaded-animation',
                '--disable-threaded-scrolling',
                '--disable-web-resources',
                '--disable-web-security',
                '--disable-xss-auditor',
                '--no-zygote',
                '--single-process',
                '--memory-pressure-off',
                '--max_old_space_size=4096'
            ],
            ignore_default_args=['--enable-automation'],
            viewport={'width': 1920, 'height': 1080},
            user_agent='Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        )
        
        self.page = self.browser.pages[0] if self.browser.pages else await self.browser.new_page()
        
        # Set additional page configurations
        await self.page.set_extra_http_headers({
            'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,image/webp,*/*;q=0.8',
            'Accept-Language': 'en-US,en;q=0.5',
            'Accept-Encoding': 'gzip, deflate, br',
            'Connection': 'keep-alive',
            'Upgrade-Insecure-Requests': '1',
        })
        
        # Set longer timeouts
        self.page.set_default_timeout(30000)
        self.page.set_default_navigation_timeout(30000)

    async def handle_popups(self):
        """Handle cookie consent and other popups with improved robustness"""
        try:
            # Wait for page to load with multiple strategies
            try:
                await self.page.wait_for_load_state('domcontentloaded', timeout=10000)
            except:
                print("DOM content loaded timeout, continuing...")
            
            # Wait a bit for any dynamic content to load
            await asyncio.sleep(2)
            
            # Handle cookie consent popup with multiple selectors
            cookie_selectors = [
                '//*[@id="didomi-notice-agree-button"]',
                'button[data-testid="cookie-accept"]',
                'button:has-text("Accept")',
                'button:has-text("Accept All")',
                'button:has-text("Agree")',
                'button:has-text("OK")',
                'button:has-text("I Accept")',
                '.cookie-accept',
                '.cookie-agree',
                '[aria-label*="Accept" i]',
                '[aria-label*="Agree" i]'
            ]
            
            for selector in cookie_selectors:
                try:
                    cookie_button = await self.page.query_selector(selector)
                    if cookie_button and await cookie_button.is_visible():
                        await cookie_button.click()
                        print(f"Cookie consent popup handled with selector: {selector}")
                        await asyncio.sleep(1)
                        break
                except Exception as e:
                    print(f"Error handling cookie popup with {selector}: {e}")
                    continue
            
            # Handle other potential popups with more comprehensive selectors
            popup_selectors = [
                'button[data-testid="close-button"]',
                '.modal-close',
                '.popup-close',
                '.close-button',
                '[aria-label="Close"]',
                '[aria-label="Dismiss"]',
                'button:has-text("Close")',
                'button:has-text("X")',
                'button:has-text("×")',
                '.close',
                '.dismiss',
                '.cancel',
                'button[class*="close"]',
                'button[class*="dismiss"]',
                '[role="button"][aria-label*="close" i]',
                '[role="button"][aria-label*="dismiss" i]'
            ]
            
            for selector in popup_selectors:
                try:
                    popup = await self.page.query_selector(selector)
                    if popup and await popup.is_visible():
                        await popup.click()
                        print(f"Closed popup with selector: {selector}")
                        await asyncio.sleep(0.5)
                except Exception as e:
                    print(f"Error handling popup with {selector}: {e}")
                    continue
                    
        except Exception as e:
            print(f"Popup handling error: {e}")
            # Continue anyway, don't let popup handling stop the process

    async def login(self):
        """Login to Robert Parker website with improved error handling and retry logic"""
        max_retries = 3
        retry_count = 0
        
        while retry_count < max_retries:
            try:
                print(f"Starting login process (attempt {retry_count + 1}/{max_retries})...")
                
                # Navigate to the homepage with better error handling
                try:
                    print("Navigating to homepage...")
                    await self.page.goto("https://www.robertparker.com/", 
                                       wait_until='domcontentloaded', 
                                       timeout=30000)
                    print("Homepage loaded successfully")
                except Exception as nav_error:
                    print(f"Navigation error: {nav_error}")
                    if "net::ERR_ABORTED" in str(nav_error) or "frame was detached" in str(nav_error):
                        print("Network error detected, retrying...")
                        retry_count += 1
                        await asyncio.sleep(2)
                        continue
                    else:
                        raise nav_error
                
                # Wait for page to be ready
                try:
                    await self.page.wait_for_load_state('networkidle', timeout=10000)
                except:
                    print("Network idle timeout, continuing anyway...")
                
                # Handle popups
                await self.handle_popups()
                
                # Check if already logged in by looking for logout button or user menu
                logout_indicators = [
                    'a[href*="logout"]',
                    '.user-menu',
                    '.account-menu',
                    '[data-testid="user-menu"]',
                    '[data-testid="account-menu"]',
                    '.user-account',
                    '.user-profile'
                ]
                
                for indicator in logout_indicators:
                    try:
                        if await self.page.query_selector(indicator):
                            print("Already logged in!")
                            return True
                    except:
                        continue
                
                print("Not logged in, attempting to login...")
                
                # Try multiple login button selectors with better error handling
                login_button_selectors = [
                    '//*[@id="root"]/header/div[1]/div/div/div[3]/div',
                    '//button[contains(text(), "Login")]',
                    '//a[contains(text(), "Login")]',
                    '//*[contains(@class, "login")]',
                    '//*[contains(@class, "signin")]',
                    '//*[contains(@class, "user")]',
                    '//*[contains(@class, "account")]'
                ]
                
                login_button = None
                for selector in login_button_selectors:
                    try:
                        login_button = await self.page.query_selector(selector)
                        if login_button and await login_button.is_visible():
                            print(f"Found login button with selector: {selector}")
                            break
                    except Exception as e:
                        print(f"Error checking selector {selector}: {e}")
                        continue
                
                if login_button:
                    print("Clicking login button...")
                    try:
                        await login_button.click()
                        await asyncio.sleep(3)
                    except Exception as click_error:
                        print(f"Error clicking login button: {click_error}")
                        retry_count += 1
                        await asyncio.sleep(2)
                        continue
                    
                    # Wait for login form to appear with multiple approaches
                    login_form_found = False
                    form_selectors = [
                        '//*[@id="user_login"]',
                        'input[type="email"]',
                        'input[name="email"]',
                        'input[type="text"][name*="email"]',
                        'input[type="text"][name*="user"]'
                    ]
                    
                    for selector in form_selectors:
                        try:
                            await self.page.wait_for_selector(selector, timeout=5000)
                            print(f"Login form appeared with selector: {selector}")
                            login_form_found = True
                            break
                        except:
                            continue
                    
                    if not login_form_found:
                        print("Login form did not appear, trying alternative approach...")
                        # Check if we're already on a login page
                        pass
                    
                    # Fill in login form with multiple selector attempts
                    email_input = None
                    password_input = None
                    submit_button = None
                    
                    # Try multiple email input selectors
                    email_selectors = [
                        '//*[@id="user_login"]',
                        'input[type="email"]',
                        'input[name="email"]',
                        'input[type="text"][name*="email"]',
                        'input[type="text"][name*="user"]',
                        'input[placeholder*="email" i]',
                        'input[placeholder*="Email" i]'
                    ]
                    
                    for selector in email_selectors:
                        try:
                            email_input = await self.page.query_selector(selector)
                            if email_input and await email_input.is_visible():
                                print(f"Found email input with selector: {selector}")
                                break
                        except:
                            continue
                    
                    # Try multiple password input selectors
                    password_selectors = [
                        '//*[@id="user_pass"]',
                        'input[type="password"]',
                        'input[name="password"]',
                        'input[name="pass"]',
                        'input[placeholder*="password" i]',
                        'input[placeholder*="Password" i]'
                    ]
                    
                    for selector in password_selectors:
                        try:
                            password_input = await self.page.query_selector(selector)
                            if password_input and await password_input.is_visible():
                                print(f"Found password input with selector: {selector}")
                                break
                        except:
                            continue
                    
                    # Try multiple submit button selectors
                    submit_selectors = [
                        '//*[@id="submit-login"]',
                        'button[type="submit"]',
                        'input[type="submit"]',
                        'button:has-text("Login")',
                        'button:has-text("Sign In")',
                        'button:has-text("Submit")',
                        '[type="submit"]'
                    ]
                    
                    for selector in submit_selectors:
                        try:
                            submit_button = await self.page.query_selector(selector)
                            if submit_button and await submit_button.is_visible():
                                print(f"Found submit button with selector: {selector}")
                                break
                        except:
                            continue
                    
                    if email_input and password_input and submit_button:
                        print("Filling login form...")
                        try:
                            await email_input.fill(self.email)
                            await password_input.fill(self.password)
                            
                            print("Submitting login form...")
                            await submit_button.click()
                            
                            # Wait for login to complete
                            await asyncio.sleep(5)
                            
                            # Verify login success with multiple indicators
                            success_indicators = [
                                'a[href*="logout"]',
                                '.user-menu',
                                '.account-menu',
                                '[data-testid="user-menu"]',
                                '[data-testid="account-menu"]',
                                '.user-account',
                                '.user-profile',
                                '.logged-in',
                                '[class*="user"]'
                            ]
                            
                            for indicator in success_indicators:
                                try:
                                    if await self.page.query_selector(indicator):
                                        print("Login successful!")
                                        return True
                                except:
                                    continue
                            
                            # Check for error messages
                            error_selectors = [
                                '.error',
                                '.alert',
                                '.message',
                                '[class*="error"]',
                                '[class*="alert"]',
                                '[class*="invalid"]',
                                '[class*="failed"]'
                            ]
                            
                            for error_selector in error_selectors:
                                try:
                                    error_element = await self.page.query_selector(error_selector)
                                    if error_element:
                                        error_text = await error_element.text_content()
                                        print(f"Login error message: {error_text}")
                                except:
                                    continue
                            
                            print("Login failed - could not find success indicators")
                            retry_count += 1
                            await asyncio.sleep(2)
                            continue
                            
                        except Exception as form_error:
                            print(f"Error filling login form: {form_error}")
                            retry_count += 1
                            await asyncio.sleep(2)
                            continue
                    else:
                        print(f"Could not find login form elements:")
                        print(f"  Email input: {'Found' if email_input else 'Not found'}")
                        print(f"  Password input: {'Found' if password_input else 'Not found'}")
                        print(f"  Submit button: {'Found' if submit_button else 'Not found'}")
                        retry_count += 1
                        await asyncio.sleep(2)
                        continue
                else:
                    print("Could not find login button with any selector")
                    retry_count += 1
                    await asyncio.sleep(2)
                    continue
                    
            except Exception as e:
                print(f"Login error (attempt {retry_count + 1}): {e}")
                retry_count += 1
                if retry_count < max_retries:
                    print(f"Retrying in 2 seconds...")
                    await asyncio.sleep(2)
                else:
                    print("Max retries reached, login failed")
                    return False
        
        print("Login failed after all retry attempts")
        return False

    async def scrape_wine_data(self, url, progress_callback=None, stop_flag=None):
        """Scrape wine data from a single URL with improved error handling"""
        async with self.semaphore:  # Limit concurrent requests
            await self.rate_limit()  # Rate limiting
            
            max_retries = 3
            retry_count = 0
            
            while retry_count < max_retries:
                if stop_flag and stop_flag():
                    if progress_callback:
                        progress_callback(f"Scraping stopped by user during: {url}")
                    return {'Full_Wine_Name': 'STOPPED', 'URL': url, 'Error': 'Scraping stopped by user'}
                try:
                    print(f"Scraping: {url} (attempt {retry_count + 1}/{max_retries})")
                    if progress_callback:
                        progress_callback(f"Scraping: {url} (attempt {retry_count + 1}/{max_retries})")
                    
                    # Navigate to the wine page with better error handling
                    try:
                        await self.page.goto(url, wait_until='domcontentloaded', timeout=30000)
                        print(f"Successfully navigated to {url}")
                    except Exception as nav_error:
                        print(f"Navigation error for {url}: {nav_error}")
                        if "net::ERR_ABORTED" in str(nav_error) or "frame was detached" in str(nav_error) or "timeout" in str(nav_error).lower():
                            retry_count += 1
                            if retry_count < max_retries:
                                print(f"Retrying navigation in 3 seconds...")
                                await asyncio.sleep(3)
                                continue
                            else:
                                print(f"Max retries reached for {url}")
                                return {
                                    'Full_Wine_Name': 'ERROR',
                                    'Producer': '',
                                    'Wine Region': '',
                                    'Color': '',
                                    'Score': '',
                                    'Drink Window': '',
                                    'Reviewed By': '',
                                    'Release Price': '',
                                    'Drink Date': '',
                                    'Tasting Note': '',
                                    'Producer Note': '',
                                    'URL': url,
                                    'Error': f"Navigation failed after {max_retries} attempts: {nav_error}"
                                }
                        else:
                            raise nav_error
                    
                    # Wait for page to be ready
                    try:
                        await self.page.wait_for_load_state('networkidle', timeout=15000)
                    except:
                        print("Network idle timeout, continuing anyway...")
                    
                    # Handle popups
                    await self.handle_popups()
                    
                    # Initialize wine data dictionary
                    wine_data = {
                        'Full_Wine_Name': '',
                        'Wine_Name': '',
                        'Vintage': '',
                        'Producer': '',
                        'Wine Region': '',
                        'Variety': '',
                        'Color': '',
                        'Score': '',
                        'Drink Window': '',
                        'Reviewed By': '',
                        'Release Price': '',
                        'Drink Date': '',
                        'Tasting Note': '',
                        'Producer Note': '',
                        'Maturity': '',
                        'Certified': '',
                        'Published Date': '',
                        'URL': url
                    }
                    
                    # Extract data using XPath selectors with better error handling
                    try:
                        # Full_Wine_Name (same XPath)
                        wine_name_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[1]/div/header/h1')
                        if wine_name_element:
                            wine_name_text = await wine_name_element.text_content()
                            wine_data['Full_Wine_Name'] = wine_name_text

                        # Producer (with fallback)
                        producer_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[1]/div/ol/li/article/div/div/div/div[1]/div[2]/span/a')
                        if not producer_element:
                            producer_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[1]/div/ol/li/article/div/div/div/div[1]/div[2]')
                        if not producer_element:
                            producer_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[1]/div/div/ol/li/article/div/div/div/div[1]/div[2]')
                        if producer_element:
                            wine_data['Producer'] = await producer_element.text_content()

                        # Extract 4-digit year (Vintage) from wine name
                        import re
                        match = re.search(r'(19|20)\d{2}', wine_name_text)
                        if match:
                            wine_data['Vintage'] = match.group(0)

                        # Remove Producer and year from Full_Wine_Name to get Wine_Name
                        wine_name_clean = wine_name_text
                        if wine_data['Producer']:
                            # Remove producer (case-insensitive, only at start)
                            wine_name_clean = re.sub(r'^' + re.escape(wine_data['Producer']) + r'\s*', '', wine_name_clean, flags=re.IGNORECASE)
                        if wine_data['Vintage']:
                            # Remove year (vintage) at the end
                            wine_name_clean = re.sub(r'\s*' + re.escape(wine_data['Vintage']) + r'$','', wine_name_clean)
                        wine_data['Wine_Name'] = wine_name_clean.strip()
                        
                        # Wine Region (with fallback)
                        region_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[1]/div/ol/li/article/div/div/div/div[2]/div[2]')
                        if not region_element:
                            region_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[1]/div/div/ol/li/article/div/div/div/div[2]')
                        if region_element:
                            region_text = await region_element.text_content()
                            # Improved regex: split before each uppercase letter that starts a word, but keep multi-word and hyphenated regions together
                            # This will match sequences like 'Southern Rhône', 'Châteauneuf-du-Pape', etc.
                            region_parts = re.findall(r'(?:[A-Z][^A-Z\s-]*(?:[\s-][A-Z][^A-Z\s-]*)*)', region_text)
                            wine_data['Wine Region'] = ', '.join([part.strip() for part in region_parts if part.strip()])
                        
                        # Color (with multiple fallback options)
                        color_element = None
                        
                        # Try multiple XPath alternatives for color
                        color_xpaths = [
                            # Original XPath
                            '//*[@id="root"]/div[1]/div/div[2]/div/div/div[1]/div/ol/li/article/div/div/div/div[4]/div[2]',
                            # Alternative with div instead of ol
                            '//*[@id="root"]/div[1]/div/div[2]/div/div/div[1]/div/div/ol/li/article/div/div/div/div[4]/div[2]',
                            # More flexible - find by position in article
                            '//article//div[4]/div[2]',
                            # Find by text content containing color-related terms
                            '//*[contains(text(), "Red") or contains(text(), "White") or contains(text(), "Rosé") or contains(text(), "Sparkling") or contains(text(), "Dessert")]',
                            # Find by class or data attributes (if available)
                            '//*[@class*="color" or @class*="type"]',
                            # Find by label text and get sibling
                            '//*[contains(text(), "Color") or contains(text(), "Type")]/following-sibling::*[1]',
                            # Generic approach - find any element with color-like text
                            '//*[matches(text(), "(Red|White|Rosé|Sparkling|Dessert|Fortified)", "i")]'
                        ]
                        
                        for xpath in color_xpaths:
                            try:
                                color_element = await self.page.query_selector(xpath)
                                if color_element:
                                    color_text = await color_element.text_content()
                                    if color_text and color_text.strip():
                                        wine_data['Color'] = color_text.strip()
                                        break
                            except Exception:
                                continue
                        
                        # If no color found, try CSS selector approach
                        if not wine_data.get('Color'):
                            try:
                                color_element = await self.page.query_selector('article div:nth-child(4) div:nth-child(2)')
                                if color_element:
                                    wine_data['Color'] = await color_element.text_content()
                            except Exception:
                                pass
                        
                        # Score
                        score_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[2]/article/div[2]/div/div/div/div[1]/div/div[2]')
                        if score_element:
                            wine_data['Score'] = await score_element.text_content()
                        
                        # Drink Window
                        drink_window_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[2]/article/div[2]/div/div/div/div[2]/div/div[3]/dl/dd')
                        if drink_window_element:
                            wine_data['Drink Window'] = await drink_window_element.text_content()
                        
                        # Reviewed By (with fallback)
                        reviewed_by_element = await self.page.query_selector('//dd/a')
                        if not reviewed_by_element:
                            reviewed_by_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[2]/article/div[2]/div/div/div/div[2]/div/div[1]/dl/dd/a')
                        if reviewed_by_element:
                            wine_data['Reviewed By'] = await reviewed_by_element.text_content()
                        
                        # Release Price
                        price_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[2]/article/div[2]/div/div/div/div[2]/div/div[2]/dl/dl/div')
                        if price_element:
                            wine_data['Release Price'] = await price_element.text_content()
                        
                        # Drink Date
                        drink_date_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[2]/article/div[2]/div/div/div/div[2]/div/div[3]/dl/dd')
                        if drink_date_element:
                            wine_data['Drink Date'] = await drink_date_element.text_content()
                        
                        # Tasting Note
                        tasting_note_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[2]/article/div[2]/div/div/div/div[2]/p[1]')
                        if tasting_note_element:
                            wine_data['Tasting Note'] = await tasting_note_element.text_content()
                        
                        # Producer Note
                        producer_note_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[2]/article/div[2]/div/div/div/div[2]/p[2]')
                        if producer_note_element:
                            wine_data['Producer Note'] = await producer_note_element.text_content()
                        
                        # Variety
                        variety_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[1]/div/ol/li/article/div/div/div/div[3]/div/a')
                        if variety_element:
                            wine_data['Variety'] = await variety_element.text_content()
                        # Maturity
                        maturity_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[1]/div/div/ol/li/article/div/div/div/div[5]')
                        if maturity_element:
                            maturity_text = await maturity_element.text_content()
                            # Check if the value starts with "Maturity:" - if not, set to "0"
                            if maturity_text and maturity_text.strip().startswith("Maturity:"):
                                wine_data['Maturity'] = maturity_text.strip()
                            else:
                                wine_data['Maturity'] = "0"
                        # Certified
                        certified_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[1]/div/div/ol/li/article/div/div/div/div[6]')
                        if certified_element:
                            certified_text = await certified_element.text_content()
                            # Check if the value contains "Certified" - if not, set to "0"
                            if certified_text and "Certified" in certified_text.strip():
                                wine_data['Certified'] = certified_text.strip()
                            else:
                                wine_data['Certified'] = "0"
                        # Published Date
                        published_date_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[2]/article/div[2]/div/div/div/p')
                        if published_date_element:
                            wine_data['Published Date'] = await published_date_element.text_content()
                        
                    except Exception as e:
                        print(f"Error extracting data from {url}: {e}")
                        retry_count += 1
                        if retry_count < max_retries:
                            print(f"Retrying data extraction in 2 seconds...")
                            await asyncio.sleep(2)
                            continue
                        else:
                            print(f"Max retries reached for data extraction from {url}")
                            return {
                                'Full_Wine_Name': 'ERROR',
                                'Producer': '',
                                'Wine Region': '',
                                'Color': '',
                                'Score': '',
                                'Drink Window': '',
                                'Reviewed By': '',
                                'Release Price': '',
                                'Drink Date': '',
                                'Tasting Note': '',
                                'Producer Note': '',
                                'URL': url,
                                'Error': f"Data extraction failed after {max_retries} attempts: {e}"
                            }
                    
                    # Clean up data (remove extra whitespace)
                    for key in wine_data:
                        if isinstance(wine_data[key], str):
                            wine_data[key] = wine_data[key].strip() if wine_data[key] else ''
                    
                    print(f"Successfully scraped: {wine_data['Full_Wine_Name']}")
                    if progress_callback:
                        progress_callback(f"Successfully scraped: {wine_data['Full_Wine_Name']}")
                    return wine_data
                    
                except Exception as e:
                    print(f"Error scraping {url} (attempt {retry_count + 1}): {e}")
                    retry_count += 1
                    if retry_count < max_retries:
                        print(f"Retrying in 3 seconds...")
                        await asyncio.sleep(3)
                    else:
                        print(f"Max retries reached for {url}")
                        if progress_callback:
                            progress_callback(f"Error scraping {url}: {e}")
                        return {
                            'Full_Wine_Name': 'ERROR',
                            'Producer': '',
                            'Wine Region': '',
                            'Color': '',
                            'Score': '',
                            'Drink Window': '',
                            'Reviewed By': '',
                            'Release Price': '',
                            'Drink Date': '',
                            'Tasting Note': '',
                            'Producer Note': '',
                            'URL': url,
                            'Error': str(e)
                        }

    async def scrape_all_wines(self, urls, progress_callback=None, stop_flag=None):
        """Scrape multiple wine URLs concurrently"""
        if not urls:
            if progress_callback:
                progress_callback("No URLs provided")
            return []
        
        # Setup browser and login
        if progress_callback:
            progress_callback("Setting up browser...")
        await self.setup_browser()
        
        if progress_callback:
            progress_callback("Logging in...")
        login_success = await self.login()
        
        if not login_success:
            if progress_callback:
                progress_callback("Failed to login. Exiting.")
            await self.browser.close()
            await self.playwright.stop()
            return []
        
        try:
            if progress_callback:
                progress_callback(f"Starting concurrent scraping of {len(urls)} URLs with max {self.max_concurrent} concurrent requests")
                progress_callback(f"Rate limit: {self.requests_per_minute} requests per minute")
            
            # Create tasks for concurrent scraping
            tasks = [self.scrape_wine_data(url, progress_callback, stop_flag) for url in urls]
            
            # Execute all tasks concurrently
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            # Filter out exceptions and get valid results
            wine_data_list = []
            for i, result in enumerate(results):
                if isinstance(result, Exception):
                    error_msg = f"Exception for URL {urls[i]}: {result}"
                    print(error_msg)
                    if progress_callback:
                        progress_callback(error_msg)
                    wine_data_list.append({
                        'Full_Wine_Name': 'ERROR',
                        'Producer': '',
                        'Wine Region': '',
                        'Color': '',
                        'Score': '',
                        'Drink Window': '',
                        'Reviewed By': '',
                        'Release Price': '',
                        'Drink Date': '',
                        'Tasting Note': '',
                        'Producer Note': '',
                        'URL': urls[i],
                        'Error': str(result)
                    })
                else:
                    wine_data_list.append(result)
            
            # Track all failed/errored/STOPPED rows for export
            if hasattr(self, 'error_rows'):
                self.error_rows.clear()
            else:
                self.error_rows = []
            for row in wine_data_list:
                if row.get('Full_Wine_Name') in ('ERROR', 'STOPPED') or row.get('Error'):
                    self.error_rows.append(row)
            return wine_data_list
            
        finally:
            await self.browser.close()
            await self.playwright.stop()

    def save_to_excel(self, wine_data_list, filename="robert_parker_wines.xlsx"):
        if not wine_data_list:
            print("No data to save")
            return
        # Deduplicate rows (ignore 'Error' field for deduplication)
        seen = set()
        deduped_list = []
        for row in wine_data_list:
            # Create a tuple of all values except 'Error' (if present)
            row_tuple = tuple((k, v) for k, v in row.items() if k != 'Error')
            if row_tuple not in seen:
                seen.add(row_tuple)
                deduped_list.append(row)
        wine_data_list = deduped_list
        try:
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Wine Data"
            # Define headers
            headers = [
                'Full_Wine_Name', 'Wine_Name', 'Vintage', 'Producer', 'Wine Region', 'Variety', 'Color', 'Score', 
                'Drink Window', 'Reviewed By', 'Release Price', 'Drink Date', 
                'Tasting Note', 'Producer Note', 'Maturity', 'Certified', 'Published Date', 'URL'
            ]
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            # Write data
            for row, wine_data in enumerate(wine_data_list, 2):
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=wine_data.get(header, ''))
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            # Save file
            wb.save(filename)
            print(f"Data saved to {filename}")
        except Exception as e:
            print(f"Error saving to Excel: {e}")

class RobertParkerGUI:
    def __init__(self, root):
        self.root = root
        self.root.title("Robert Parker Wine Scraper (Concurrent)")
        self.root.geometry("800x700")
        
        # Variables
        self.email_var = tk.StringVar(value="david@domaine.com.tw")
        self.password_var = tk.StringVar(value="Dwc123")
        self.output_filename_var = tk.StringVar(value="robert_parker_wines.xlsx")
        self.max_concurrent_var = tk.IntVar(value=5)
        self.requests_per_minute_var = tk.IntVar(value=30)
        self.is_scraping = False
        self.stop_scraping = False
        
        # Speed tracking variables
        self.start_time = None
        self.completed_requests = 0
        self.speed_var = tk.StringVar(value="Speed: 0 requests/min")
        self.speed_update_timer = None
        
        # Time tracking variables
        self.scraping_start_time = None
        self.time_var = tk.StringVar(value="Time: 00:00:00")
        self.time_update_timer = None
        
        # URL tracking variables
        self.current_url_index = 0
        self.total_urls = 0
        self.url_progress_var = tk.StringVar(value="URLs: 0/0")
        
        self.error_logs = []  # Track error logs for export
        self.setup_ui()
        
    def setup_ui(self):
        # Main frame
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="Robert Parker Wine Scraper", font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))
        
        # Login credentials
        ttk.Label(main_frame, text="Email:").grid(row=1, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.email_var, width=40).grid(row=1, column=1, sticky=(tk.W, tk.E), pady=5)
        
        ttk.Label(main_frame, text="Password:").grid(row=2, column=0, sticky=tk.W, pady=5)
        ttk.Entry(main_frame, textvariable=self.password_var, show="*", width=40).grid(row=2, column=1, sticky=(tk.W, tk.E), pady=5)
        
        # Performance settings
        perf_frame = ttk.LabelFrame(main_frame, text="Performance Settings", padding="10")
        perf_frame.grid(row=3, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        perf_frame.columnconfigure(1, weight=1)
        
        ttk.Label(perf_frame, text="Max Concurrent Requests:").grid(row=0, column=0, sticky=tk.W, pady=5)
        concurrent_spinbox = ttk.Spinbox(perf_frame, from_=1, to=20, textvariable=self.max_concurrent_var, width=10)
        concurrent_spinbox.grid(row=0, column=1, sticky=tk.W, pady=5)
        
        ttk.Label(perf_frame, text="Requests per Minute:").grid(row=1, column=0, sticky=tk.W, pady=5)
        rate_spinbox = ttk.Spinbox(perf_frame, from_=10, to=100, textvariable=self.requests_per_minute_var, width=10)
        rate_spinbox.grid(row=1, column=1, sticky=tk.W, pady=5)
        
        # URLs input
        ttk.Label(main_frame, text="Wine URLs (one per line):").grid(row=4, column=0, sticky=tk.W, pady=(10, 5))
        
        url_frame = ttk.Frame(main_frame)
        url_frame.grid(row=5, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        url_frame.columnconfigure(0, weight=1)
        url_frame.rowconfigure(0, weight=1)
        
        self.url_text = scrolledtext.ScrolledText(url_frame, height=8, width=80)
        self.url_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Sample URL
        sample_url = "https://www.robertparker.com/wines/xiPRuQod7Qy2rC5bv/louis-jadot-chassagne-montrachet-1er-cru-morgeot-maison-louis-jadot-1985"
        self.url_text.insert(tk.END, sample_url)
        
        # Output filename
        ttk.Label(main_frame, text="Output Filename:").grid(row=6, column=0, sticky=tk.W, pady=(10, 5))
        filename_frame = ttk.Frame(main_frame)
        filename_frame.grid(row=7, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=5)
        filename_frame.columnconfigure(0, weight=1)
        
        ttk.Entry(filename_frame, textvariable=self.output_filename_var).grid(row=0, column=0, sticky=(tk.W, tk.E))
        ttk.Button(filename_frame, text="Browse", command=self.browse_filename).grid(row=0, column=1, padx=(5, 0))
        
        # Control buttons
        button_frame = ttk.Frame(main_frame)
        button_frame.grid(row=8, column=0, columnspan=2, pady=20)
        
        self.start_button = ttk.Button(button_frame, text="Start Scraping", command=self.start_scraping)
        self.start_button.pack(side=tk.LEFT, padx=5)
        
        self.stop_button = ttk.Button(button_frame, text="Stop Scraping", command=self.stop_scraping_func, state=tk.DISABLED)
        self.stop_button.pack(side=tk.LEFT, padx=5)
        
        # Progress bar and speed display
        progress_frame = ttk.Frame(main_frame)
        progress_frame.grid(row=9, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=10)
        progress_frame.columnconfigure(0, weight=1)
        
        self.progress_var = tk.DoubleVar()
        self.progress_bar = ttk.Progressbar(progress_frame, variable=self.progress_var, maximum=100)
        self.progress_bar.grid(row=0, column=0, sticky=(tk.W, tk.E), pady=(0, 5))
        
        # Speed display
        self.speed_label = ttk.Label(progress_frame, textvariable=self.speed_var, font=("Arial", 10, "bold"))
        self.speed_label.grid(row=1, column=0, sticky=tk.W)
        
        # Time display
        self.time_label = ttk.Label(progress_frame, textvariable=self.time_var, font=("Arial", 10, "bold"))
        self.time_label.grid(row=1, column=1, sticky=tk.E)
        
        # URL progress display
        self.url_progress_label = ttk.Label(progress_frame, textvariable=self.url_progress_var, font=("Arial", 10, "bold"))
        self.url_progress_label.grid(row=2, column=0, columnspan=2, sticky=tk.W, pady=(5, 0))
        
        # Log output
        ttk.Label(main_frame, text="Log Output:").grid(row=10, column=0, sticky=tk.W, pady=(10, 5))
        
        log_frame = ttk.Frame(main_frame)
        log_frame.grid(row=11, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=5)
        log_frame.columnconfigure(0, weight=1)
        log_frame.rowconfigure(0, weight=1)
        
        self.log_text = scrolledtext.ScrolledText(log_frame, height=10, width=80)
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Add Export Error Log button below log output
        export_error_btn = ttk.Button(self.root, text="Export Error Log", command=self.export_error_log)
        export_error_btn.place(x=650, y=660)  # Adjust position as needed
        
        # Configure main frame row weights
        main_frame.rowconfigure(5, weight=1)
        main_frame.rowconfigure(11, weight=1)
        
    def browse_filename(self):
        filename = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")]
        )
        if filename:
            self.output_filename_var.set(filename)
    
    def log_message(self, message):
        """Add message to log output"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        self.log_text.insert(tk.END, f"[{timestamp}] {message}\n")
        self.log_text.see(tk.END)
        self.root.update_idletasks()
    
    def progress_callback(self, message):
        """Custom progress callback that logs and updates speed, and collects errors"""
        self.log_message(message)
        # Check if this is a completion message
        if "Successfully scraped:" in message or "Error scraping" in message:
            self.increment_completed_requests()
            self.update_url_progress()
        # Collect error logs
        if message.startswith("Error scraping") or message.startswith("Exception for URL"):
            self.error_logs.append(message)

    def update_speed(self):
        """Update the speed display"""
        if self.start_time and self.completed_requests > 0:
            elapsed_time = time.time() - self.start_time
            if elapsed_time > 0:
                requests_per_minute = (self.completed_requests / elapsed_time) * 60
                self.speed_var.set(f"Speed: {requests_per_minute:.1f} requests/min")
            else:
                self.speed_var.set("Speed: Calculating...")
        else:
            self.speed_var.set("Speed: 0 requests/min")
        
        # Update time display
        self.update_time()
        
        # Schedule next update if scraping is active
        if self.is_scraping:
            self.speed_update_timer = self.root.after(1000, self.update_speed)  # Update every second
    
    def update_time(self):
        """Update the time display"""
        if self.scraping_start_time:
            elapsed_time = time.time() - self.scraping_start_time
            hours = int(elapsed_time // 3600)
            minutes = int((elapsed_time % 3600) // 60)
            seconds = int(elapsed_time % 60)
            self.time_var.set(f"Time: {hours:02d}:{minutes:02d}:{seconds:02d}")
        else:
            self.time_var.set("Time: 00:00:00")
    
    def increment_completed_requests(self):
        """Increment completed requests counter and update speed"""
        self.completed_requests += 1
        self.update_speed()
    
    def update_url_progress(self):
        """Update the URL progress display"""
        if self.total_urls > 0:
            self.current_url_index = min(self.completed_requests, self.total_urls)
            self.url_progress_var.set(f"URLs: {self.current_url_index}/{self.total_urls}")
        else:
            self.url_progress_var.set("URLs: 0/0")
    
    def get_urls(self):
        """Get URLs from text input"""
        urls_text = self.url_text.get(1.0, tk.END).strip()
        if not urls_text:
            return []
        
        urls = [url.strip() for url in urls_text.split('\n') if url.strip()]
        return urls
    
    def start_scraping(self):
        """Start the scraping process"""
        if self.is_scraping:
            return
        
        urls = self.get_urls()
        if not urls:
            messagebox.showerror("Error", "Please enter at least one URL")
            return
        
        self.is_scraping = True
        self.stop_scraping = False
        self.start_button.config(state=tk.DISABLED)
        self.stop_button.config(state=tk.NORMAL)
        
        # Clear log and reset counters
        self.log_text.delete(1.0, tk.END)
        self.progress_var.set(0)
        self.start_time = time.time()
        self.completed_requests = 0
        self.speed_var.set("Speed: 0 requests/min")
        self.scraping_start_time = time.time() # Set scraping start time
        
        # Initialize URL progress tracking
        self.total_urls = len(urls)
        self.current_url_index = 0
        self.url_progress_var.set(f"URLs: 0/{self.total_urls}")
        
        # Start speed updates
        self.update_speed()
        
        # Start scraping in a separate thread
        scraping_thread = threading.Thread(target=self.run_scraping, args=(urls,))
        scraping_thread.daemon = True
        scraping_thread.start()
    
    def stop_scraping_func(self):
        """Stop the scraping process"""
        self.stop_scraping = True
        self.log_message("Stopping scraping...")
    
    def run_scraping(self, urls):
        """Run the scraping process in a separate thread"""
        try:
            # Create new event loop for this thread
            loop = asyncio.new_event_loop()
            asyncio.set_event_loop(loop)
            
            # Run the scraping
            loop.run_until_complete(self.scrape_all_wines(urls, stop_flag=lambda: self.stop_scraping))
            
        except Exception as e:
            self.log_message(f"Error in scraping thread: {e}")
        finally:
            # Reset UI
            self.root.after(0, self.reset_ui)
    
    def reset_ui(self):
        """Reset UI after scraping completes"""
        self.is_scraping = False
        self.start_button.config(state=tk.NORMAL)
        self.stop_button.config(state=tk.DISABLED)
        self.progress_var.set(100)
        
        # Cancel speed update timer
        if self.speed_update_timer:
            self.root.after_cancel(self.speed_update_timer)
            self.speed_update_timer = None
    
    async def scrape_all_wines(self, urls, progress_callback=None, stop_flag=None):
        """Scrape all wines asynchronously"""
        try:
            self.log_message(f"Starting scraping of {len(urls)} URLs")
            self.log_message(f"Max concurrent requests: {self.max_concurrent_var.get()}")
            self.log_message(f"Rate limit: {self.requests_per_minute_var.get()} requests per minute")
            
            # Create scraper instance
            scraper = RobertParkerScraper(
                email=self.email_var.get(),
                password=self.password_var.get(),
                max_concurrent=self.max_concurrent_var.get(),
                requests_per_minute=self.requests_per_minute_var.get()
            )
            
            # Start timing
            start_time = time.time()
            
            # Scrape all wines
            wine_data_list = await scraper.scrape_all_wines(urls, self.progress_callback, stop_flag)
            self.last_wine_data_list = wine_data_list  # Store for error export
            
            # End timing
            end_time = time.time()
            elapsed_time = end_time - start_time
            
            self.log_message(f"Scraping completed in {elapsed_time:.2f} seconds")
            self.log_message(f"Average time per URL: {elapsed_time/len(urls):.2f} seconds")
            
            # Save results
            if wine_data_list:
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                filename = f"robert_parker_wines_{timestamp}.xlsx"
                scraper.save_to_excel(wine_data_list, filename)
                
                # Print summary
                successful = sum(1 for wine in wine_data_list if wine.get('Full_Wine_Name') != 'ERROR')
                self.log_message(f"Summary:")
                self.log_message(f"Total URLs processed: {len(urls)}")
                self.log_message(f"Successful scrapes: {successful}")
                self.log_message(f"Failed scrapes: {len(urls) - successful}")
                self.log_message(f"Results saved to: {filename}")
                
                messagebox.showinfo("Success", f"Scraping completed!\nResults saved to: {filename}")
            else:
                self.log_message("No data was scraped")
                messagebox.showwarning("Warning", "No data was scraped")
                
        except Exception as e:
            self.log_message(f"Error during scraping: {e}")
            messagebox.showerror("Error", f"An error occurred during scraping: {e}")

    def export_error_log(self):
        """Export all failed data rows (with 'ERROR', 'STOPPED', or non-empty 'Error' field) to a CSV file"""
        error_rows = getattr(self, 'error_rows', None)
        if not error_rows:
            messagebox.showinfo("No Errors", "There are no error rows to export.")
            return
        file_path = filedialog.asksaveasfilename(
            defaultextension=".csv",
            filetypes=[("CSV files", "*.csv"), ("Text files", "*.txt"), ("All files", "*.*")],
            title="Save Error Log As"
        )
        if not file_path:
            return
        try:
            import csv
            with open(file_path, "w", encoding="utf-8", newline='') as f:
                writer = csv.DictWriter(f, fieldnames=error_rows[0].keys())
                writer.writeheader()
                for row in error_rows:
                    writer.writerow(row)
            messagebox.showinfo("Export Successful", f"Error log exported to: {file_path}")
        except Exception as e:
            messagebox.showerror("Export Failed", f"Failed to export error log: {e}")

def main():
    root = tk.Tk()
    app = RobertParkerGUI(root)
    root.mainloop()

if __name__ == "__main__":
    main() 