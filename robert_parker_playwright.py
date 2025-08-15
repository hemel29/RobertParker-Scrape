import asyncio
import json
import os
from datetime import datetime
from playwright.async_api import async_playwright
import openpyxl
from openpyxl.styles import Font
import time
import aiohttp
from asyncio import Semaphore

class RobertParkerScraper:
    def __init__(self, email, password, max_concurrent=5, requests_per_minute=30):
        self.email = email
        self.password = password
        self.cookies_file = "robert_parker_cookies.json"
        self.max_concurrent = max_concurrent
        self.requests_per_minute = requests_per_minute
        self.semaphore = Semaphore(max_concurrent)
        self.rate_limit_delay = 60.0 / requests_per_minute
        self.last_request_time = 0
        
    async def rate_limit(self):
        """Ensure we don't exceed the rate limit"""
        current_time = time.time()
        time_since_last = current_time - self.last_request_time
        if time_since_last < self.rate_limit_delay:
            await asyncio.sleep(self.rate_limit_delay - time_since_last)
        self.last_request_time = time.time()

    async def setup_browser(self):
        """Set up Playwright browser with persistent context"""
        self.playwright = await async_playwright().start()
        
        # Use persistent context for better performance and cookie management
        user_data_dir = os.path.join(os.getcwd(), "browser_data")
        self.browser = await self.playwright.chromium.launch_persistent_context(
            user_data_dir=user_data_dir,
            headless=False,  # Set to True for production
            args=[
                '--no-sandbox',
                '--disable-dev-shm-usage',
                '--disable-gpu',
                '--disable-web-security',
                '--disable-features=VizDisplayCompositor'
            ]
        )
        
        self.page = self.browser.pages[0] if self.browser.pages else await self.browser.new_page()
        
        # Set viewport and user agent
        await self.page.set_viewport_size({"width": 1920, "height": 1080})
        await self.page.set_extra_http_headers({
            'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36'
        })

    async def handle_popups(self):
        """Handle cookie consent and other popups"""
        try:
            # Wait for page to load
            await self.page.wait_for_load_state('networkidle', timeout=10000)
            
            # Handle cookie consent popup with the specific XPath
            cookie_button = await self.page.query_selector('//*[@id="didomi-notice-agree-button"]')
            if cookie_button:
                await cookie_button.click()
                print("Cookie consent popup handled")
                await asyncio.sleep(1)
            
            # Handle other potential popups
            popup_selectors = [
                'button[data-testid="close-button"]',
                '.modal-close',
                '.popup-close',
                '[aria-label="Close"]',
                '.close-button'
            ]
            
            for selector in popup_selectors:
                try:
                    popup = await self.page.query_selector(selector)
                    if popup and await popup.is_visible():
                        await popup.click()
                        print(f"Closed popup with selector: {selector}")
                        await asyncio.sleep(0.5)
                except:
                    continue
                    
        except Exception as e:
            print(f"Popup handling error: {e}")

    async def login(self):
        """Login to Robert Parker website"""
        try:
            print("Starting login process...")
            
            # Check if we're already logged in
            await self.page.goto("https://www.robertparker.com/", wait_until='networkidle', timeout=15000)
            await self.handle_popups()
            
            # Check if already logged in by looking for logout button or user menu
            logout_indicators = [
                'a[href*="logout"]',
                '.user-menu',
                '.account-menu',
                '[data-testid="user-menu"]',
                '[data-testid="account-menu"]',
                '.user-account'
            ]
            
            for indicator in logout_indicators:
                if await self.page.query_selector(indicator):
                    print("Already logged in!")
                    return True
            
            print("Not logged in, attempting to login...")
            
            # Try multiple login button selectors
            login_button_selectors = [
                '//*[@id="root"]/header/div[1]/div/div/div[3]/div',
                '//button[contains(text(), "Login")]',
                '//a[contains(text(), "Login")]',
                '//*[contains(@class, "login")]',
                '//*[contains(@class, "signin")]'
            ]
            
            login_button = None
            for selector in login_button_selectors:
                try:
                    login_button = await self.page.query_selector(selector)
                    if login_button:
                        print(f"Found login button with selector: {selector}")
                        break
                except:
                    continue
            
            if login_button:
                print("Clicking login button...")
                await login_button.click()
                await asyncio.sleep(3)
                
                # Wait for login form to appear
                try:
                    await self.page.wait_for_selector('//*[@id="user_login"]', timeout=10000)
                    print("Login form appeared")
                except:
                    print("Login form did not appear, trying alternative approach...")
                    # Try to find login form on current page
                    pass
                
                # Fill in login form
                email_input = await self.page.query_selector('//*[@id="user_login"]')
                password_input = await self.page.query_selector('//*[@id="user_pass"]')
                submit_button = await self.page.query_selector('//*[@id="submit-login"]')
                
                if not email_input:
                    print("Could not find email input, trying alternative selectors...")
                    email_input = await self.page.query_selector('input[type="email"]') or await self.page.query_selector('input[name="email"]')
                
                if not password_input:
                    print("Could not find password input, trying alternative selectors...")
                    password_input = await self.page.query_selector('input[type="password"]') or await self.page.query_selector('input[name="password"]')
                
                if not submit_button:
                    print("Could not find submit button, trying alternative selectors...")
                    submit_button = await self.page.query_selector('button[type="submit"]') or await self.page.query_selector('input[type="submit"]')
                
                if email_input and password_input and submit_button:
                    print("Filling login form...")
                    await email_input.fill(self.email)
                    await password_input.fill(self.password)
                    
                    print("Submitting login form...")
                    await submit_button.click()
                    
                    # Wait for login to complete
                    await asyncio.sleep(5)
                    
                    # Verify login success with multiple indicators
                    success_indicators = [
                        'a[href*="logout"]',
                        '.user-menu',
                        '.account-menu',
                        '[data-testid="user-menu"]',
                        '[data-testid="account-menu"]',
                        '.user-account',
                        '.user-profile'
                    ]
                    
                    for indicator in success_indicators:
                        if await self.page.query_selector(indicator):
                            print("Login successful!")
                            return True
                    
                    # Check for error messages
                    error_selectors = [
                        '.error',
                        '.alert',
                        '.message',
                        '[class*="error"]',
                        '[class*="alert"]'
                    ]
                    
                    for error_selector in error_selectors:
                        error_element = await self.page.query_selector(error_selector)
                        if error_element:
                            error_text = await error_element.text_content()
                            print(f"Login error message: {error_text}")
                    
                    print("Login failed - could not find success indicators")
                    return False
                else:
                    print(f"Could not find login form elements:")
                    print(f"  Email input: {'Found' if email_input else 'Not found'}")
                    print(f"  Password input: {'Found' if password_input else 'Not found'}")
                    print(f"  Submit button: {'Found' if submit_button else 'Not found'}")
                    return False
            else:
                print("Could not find login button with any selector")
                return False
                
        except Exception as e:
            print(f"Login error: {e}")
            return False

    async def scrape_wine_data(self, url):
        """Scrape wine data from a single URL"""
        async with self.semaphore:  # Limit concurrent requests
            await self.rate_limit()  # Rate limiting
            
            try:
                print(f"Scraping: {url}")
                
                # Navigate to the wine page
                await self.page.goto(url, wait_until='networkidle', timeout=10000)
                await self.handle_popups()
                
                # Initialize wine data dictionary
                wine_data = {
                    'Wine Name': '',
                    'Producer': '',
                    'Wine Region': '',
                    'Color': '',
                    'Score': '',
                    'Drink Window': '',
                    'Reviewed By': '',
                    'Release Price': '',
                    'Drink Date': '',
                    'Tasting Note': '',
                    'Producer Note': '',
                    'URL': url
                }
                
                # Extract data using XPath selectors
                try:
                    # Wine Name
                    wine_name_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[1]/div/header/h1')
                    if wine_name_element:
                        wine_data['Wine Name'] = await wine_name_element.text_content()
                    
                    # Producer (with fallback)
                    producer_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[1]/div/ol/li/article/div/div/div/div[1]/div[2]/span/a')
                    if not producer_element:
                        producer_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[1]/div/ol/li/article/div/div/div/div[1]/div[2]')
                    if not producer_element:
                        producer_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[1]/div/div/ol/li/article/div/div/div/div[1]/div[2]')
                    if producer_element:
                        wine_data['Producer'] = await producer_element.text_content()
                    
                    # Wine Region (with fallback)
                    region_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[1]/div/ol/li/article/div/div/div/div[2]/div[2]')
                    if not region_element:
                        region_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[1]/div/div/ol/li/article/div/div/div/div[2]')
                    if region_element:
                        wine_data['Wine Region'] = await region_element.text_content()
                    
                    # Color (with multiple fallback options)
                    color_element = None
                    
                    # Try multiple XPath alternatives for color
                    color_xpaths = [
                        # Original XPath
                        '//*[@id="root"]/div[1]/div/div[2]/div/div/div[1]/div/ol/li/article/div/div/div/div[4]/div[2]',
                        # Alternative with div instead of ol
                        '//*[@id="root"]/div[1]/div/div[2]/div/div/div[1]/div/div/ol/li/article/div/div/div/div[4]/div[2]',
                        # More flexible - find by position in article
                        '//article//div[4]/div[2]',
                        # Find by text content containing color-related terms
                        '//*[contains(text(), "Red") or contains(text(), "White") or contains(text(), "Rosé") or contains(text(), "Sparkling") or contains(text(), "Dessert")]',
                        # Find by class or data attributes (if available)
                        '//*[@class*="color" or @class*="type"]',
                        # Find by label text and get sibling
                        '//*[contains(text(), "Color") or contains(text(), "Type")]/following-sibling::*[1]',
                        # Generic approach - find any element with color-like text
                        '//*[matches(text(), "(Red|White|Rosé|Sparkling|Dessert|Fortified)", "i")]'
                    ]
                    
                    for xpath in color_xpaths:
                        try:
                            color_element = await self.page.query_selector(xpath)
                            if color_element:
                                color_text = await color_element.text_content()
                                if color_text and color_text.strip():
                                    wine_data['Color'] = color_text.strip()
                                    break
                        except Exception:
                            continue
                    
                    # If no color found, try CSS selector approach
                    if not wine_data.get('Color'):
                        try:
                            color_element = await self.page.query_selector('article div:nth-child(4) div:nth-child(2)')
                            if color_element:
                                wine_data['Color'] = await color_element.text_content()
                        except Exception:
                            pass
                    
                    # Score
                    score_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[2]/article/div[2]/div/div/div/div[1]/div/div[2]')
                    if score_element:
                        wine_data['Score'] = await score_element.text_content()
                    
                    # Drink Window
                    drink_window_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[2]/article/div[2]/div/div/div/div[2]/div/div[3]/dl/dd')
                    if drink_window_element:
                        wine_data['Drink Window'] = await drink_window_element.text_content()
                    
                    # Reviewed By (with fallback)
                    reviewed_by_element = await self.page.query_selector('//dd/a')
                    if not reviewed_by_element:
                        reviewed_by_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[2]/article/div[2]/div/div/div/div[2]/div/div[1]/dl/dd/a')
                    if reviewed_by_element:
                        wine_data['Reviewed By'] = await reviewed_by_element.text_content()
                    
                    # Release Price
                    price_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[2]/article/div[2]/div/div/div/div[2]/div/div[2]/dl/dl/div')
                    if price_element:
                        wine_data['Release Price'] = await price_element.text_content()
                    
                    # Drink Date
                    drink_date_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[2]/article/div[2]/div/div/div/div[2]/div/div[3]/dl/dd')
                    if drink_date_element:
                        wine_data['Drink Date'] = await drink_date_element.text_content()
                    
                    # Tasting Note
                    tasting_note_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[2]/article/div[2]/div/div/div/div[2]/p[1]')
                    if tasting_note_element:
                        wine_data['Tasting Note'] = await tasting_note_element.text_content()
                    
                    # Producer Note
                    producer_note_element = await self.page.query_selector('//*[@id="root"]/div[1]/div/div[2]/div/div/div[2]/article/div[2]/div/div/div/div[2]/p[2]')
                    if producer_note_element:
                        wine_data['Producer Note'] = await producer_note_element.text_content()
                    
                except Exception as e:
                    print(f"Error extracting data from {url}: {e}")
                
                # Clean up data (remove extra whitespace)
                for key in wine_data:
                    if isinstance(wine_data[key], str):
                        wine_data[key] = wine_data[key].strip() if wine_data[key] else ''
                
                print(f"Successfully scraped: {wine_data['Wine Name']}")
                return wine_data
                
            except Exception as e:
                print(f"Error scraping {url}: {e}")
                return {
                    'Wine Name': 'ERROR',
                    'Producer': '',
                    'Wine Region': '',
                    'Color': '',
                    'Score': '',
                    'Drink Window': '',
                    'Reviewed By': '',
                    'Release Price': '',
                    'Drink Date': '',
                    'Tasting Note': '',
                    'Producer Note': '',
                    'URL': url,
                    'Error': str(e)
                }

    async def scrape_all_wines(self, urls):
        """Scrape multiple wine URLs concurrently"""
        if not urls:
            print("No URLs provided")
            return []
        
        # Setup browser and login
        await self.setup_browser()
        login_success = await self.login()
        
        if not login_success:
            print("Failed to login. Exiting.")
            await self.browser.close()
            await self.playwright.stop()
            return []
        
        try:
            print(f"Starting concurrent scraping of {len(urls)} URLs with max {self.max_concurrent} concurrent requests")
            print(f"Rate limit: {self.requests_per_minute} requests per minute")
            
            # Create tasks for concurrent scraping
            tasks = [self.scrape_wine_data(url) for url in urls]
            
            # Execute all tasks concurrently
            results = await asyncio.gather(*tasks, return_exceptions=True)
            
            # Filter out exceptions and get valid results
            wine_data_list = []
            for i, result in enumerate(results):
                if isinstance(result, Exception):
                    print(f"Exception for URL {urls[i]}: {result}")
                    wine_data_list.append({
                        'Wine Name': 'ERROR',
                        'Producer': '',
                        'Wine Region': '',
                        'Color': '',
                        'Score': '',
                        'Drink Window': '',
                        'Reviewed By': '',
                        'Release Price': '',
                        'Drink Date': '',
                        'Tasting Note': '',
                        'Producer Note': '',
                        'URL': urls[i],
                        'Error': str(result)
                    })
                else:
                    wine_data_list.append(result)
            
            return wine_data_list
            
        finally:
            await self.browser.close()
            await self.playwright.stop()

    def save_to_excel(self, wine_data_list, filename="robert_parker_wines.xlsx"):
        """Save scraped data to Excel file"""
        if not wine_data_list:
            print("No data to save")
            return
        
        try:
            # Create workbook and worksheet
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Wine Data"
            
            # Define headers
            headers = [
                'Wine Name', 'Producer', 'Wine Region', 'Color', 'Score', 
                'Drink Window', 'Reviewed By', 'Release Price', 'Drink Date', 
                'Tasting Note', 'Producer Note', 'URL'
            ]
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = Font(bold=True)
            
            # Write data
            for row, wine_data in enumerate(wine_data_list, 2):
                for col, header in enumerate(headers, 1):
                    ws.cell(row=row, column=col, value=wine_data.get(header, ''))
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save file
            wb.save(filename)
            print(f"Data saved to {filename}")
            
        except Exception as e:
            print(f"Error saving to Excel: {e}")

async def main():
    # Configuration
    email = "david@domaine.com.tw"
    password = "Dwc123"
    
    # Performance settings
    max_concurrent = 5  # Number of concurrent requests
    requests_per_minute = 30  # Rate limit
    
    # Sample URLs - replace with your actual URLs
    urls = [
        "https://www.robertparker.com/wines/xiPRuQod7Qy2rC5bv/louis-jadot-chassagne-montrachet-1er-cru-morgeot-maison-louis-jadot-1985"
    ]
    
    # Add more URLs here for testing
    # urls.extend([
    #     "https://www.robertparker.com/wines/another-wine-url",
    #     "https://www.robertparker.com/wines/another-wine-url-2"
    # ])
    
    print("=== Robert Parker Wine Scraper (Concurrent Version) ===")
    print(f"Max concurrent requests: {max_concurrent}")
    print(f"Rate limit: {requests_per_minute} requests per minute")
    print(f"Total URLs to scrape: {len(urls)}")
    print()
    
    # Create scraper instance
    scraper = RobertParkerScraper(
        email=email, 
        password=password,
        max_concurrent=max_concurrent,
        requests_per_minute=requests_per_minute
    )
    
    # Start timing
    start_time = time.time()
    
    # Scrape all wines
    wine_data_list = await scraper.scrape_all_wines(urls)
    
    # End timing
    end_time = time.time()
    elapsed_time = end_time - start_time
    
    print(f"\nScraping completed in {elapsed_time:.2f} seconds")
    print(f"Average time per URL: {elapsed_time/len(urls):.2f} seconds")
    
    # Save results
    if wine_data_list:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"robert_parker_wines_{timestamp}.xlsx"
        scraper.save_to_excel(wine_data_list, filename)
        
        # Print summary
        successful = sum(1 for wine in wine_data_list if wine.get('Wine Name') != 'ERROR')
        print(f"\nSummary:")
        print(f"Total URLs processed: {len(urls)}")
        print(f"Successful scrapes: {successful}")
        print(f"Failed scrapes: {len(urls) - successful}")
        print(f"Results saved to: {filename}")

if __name__ == "__main__":
    asyncio.run(main()) 